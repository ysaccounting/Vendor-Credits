"""
Vendor Credits — backend.

Upload a QuickBooks **Balance Sheet** export and an **A/P Aging Summary** export
(.xlsx / .xlsm / .csv). Produces a workbook with three tabs — Vendor Credits,
Other TC, Other AP — reproducing the Vendor_Credits.xlsm Power Query workbook.

Ticket credits are read only from the **Prepaid Inventory** (a.k.a. "Ticket Credits")
section of the Balance Sheet. The **Missing Buy Ins** and **Deposits** subgroups are
excluded. League/group headers (MLB, NCAA > Basketball, Concerts, …) are carried into
the output as bold separator rows; the team/venue accounts (those ending in "(TC)")
are matched against the A/P Aging.

Matching tolerates formatting differences between the two reports (case, spacing,
punctuation, St/Saint, city abbreviations such as LA/Los Angeles, soccer SC/FC/CF in
any position) while preserving each side's exact name in the output. A sport word
(Football/Basketball/…) must agree when present on both sides. A/P vendors ending in a
PSL marker are never matched. A maintainable ALIASES list forces specific pairs.
"""

import io
import os
import csv
import re
import time
import uuid
import shutil
import calendar
import tempfile
import datetime as dt

from flask import Flask, request, jsonify, send_file, send_from_directory, abort
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder=None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STORE_DIR = os.path.join(tempfile.gettempdir(), "vcred_store")
os.makedirs(STORE_DIR, exist_ok=True)


# =========================================================================== #
# MATCHING CONFIG — edit these to tune how Balance Sheet names line up with A/P
# =========================================================================== #

# Force two names (one as it appears on either report) to be treated as the same
# vendor. Wins over every heuristic. Order within a pair does not matter.
ALIASES = [
    ("Milwakee Bucks", "Milwaukee Bucks"),      # example: source typo -> correct
    # ("TC or AP name", "the other report's name"),
]

# Whole-token city abbreviation expansions (applied to both sides before matching).
CITY_ABBREV = {
    "la": "los angeles", "ny": "new york", "nyc": "new york",
    "sf": "san francisco", "sd": "san diego", "kc": "kansas city",
    "tb": "tampa bay", "gb": "green bay", "no": "new orleans",
    "ne": "new england", "stl": "saint louis",
    "st": "saint", "ft": "fort", "mt": "mount",
}

SPORTS = {"football", "basketball", "baseball", "hockey", "soccer"}
SOCCER_SUFFIX = {"fc", "sc", "cf"}      # dropped + position-independent
EXCLUDED_GROUPS = {"missing buy ins", "deposits"}
PARENT_NAMES = {"prepaid inventory", "ticket credits"}


# =========================================================================== #
# Generic cell / file helpers
# =========================================================================== #

_NUM_FORMULA = re.compile(r"-?\d+(\.\d+)?")


def _cleanup_old(max_age_seconds=12 * 3600):
    now = time.time()
    for name in os.listdir(STORE_DIR):
        path = os.path.join(STORE_DIR, name)
        try:
            if os.path.isdir(path) and now - os.path.getmtime(path) > max_age_seconds:
                shutil.rmtree(path, ignore_errors=True)
        except OSError:
            pass


def _amount(v):
    """Coerce a cell to float. Handles numbers, '=123.45' literal formulas,
    '$1,234.56', '(123)' negatives. Cell-ref formulas (=B6+C6) -> None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", " "):
        return None
    if s.startswith("="):
        body = s[1:]
        return float(body) if _NUM_FORMULA.fullmatch(body) else None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", "."):
        return None
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None


def _rows_from_upload(filename, data):
    low = filename.lower()
    if low.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return [list(r) for r in csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    best, best_score = None, -1
    for ws in wb.worksheets:
        score = ws.max_row * (ws.max_column or 1)
        if score > best_score:
            best, best_score = ws, score
    rows = [list(r) for r in best.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_balance_sheet(filename, data):
    """Like _rows_from_upload, but bakes each row's indentation into leading spaces
    on the account-name cell. QuickBooks balance sheets indicate the account tree
    either with literal leading spaces OR with Excel's cell indent property; this
    captures both so the Prepaid Inventory section can be scoped by depth."""
    if filename.lower().endswith(".csv"):
        return _rows_from_upload(filename, data)
    wb = load_workbook(io.BytesIO(data), data_only=False, read_only=False)
    best, best_score = None, -1
    for ws in wb.worksheets:
        score = ws.max_row * (ws.max_column or 1)
        if score > best_score:
            best, best_score = ws, score
    out = []
    for row in best.iter_rows():
        name_cell = next((c for c in row if isinstance(c.value, str) and c.value.strip()),
                         None)
        vals = []
        for c in row:
            v = c.value
            if c is name_cell:
                raw = str(v)
                lead = len(raw) - len(raw.lstrip())
                try:
                    align = int(round(c.alignment.indent or 0))
                except (TypeError, ValueError):
                    align = 0
                v = (" " * (lead + align * 3)) + raw.strip()
            vals.append(v)
        out.append(vals)
    wb.close()
    return out


def _find_asof(rows):
    for row in rows[:10]:
        for cell in row[:4]:
            if isinstance(cell, str) and "as of" in cell.lower():
                m = re.search(r"as of\s+(.*)$", cell, re.I)
                if m:
                    for fmt in ("%B %d, %Y", "%b %d, %Y", "%m/%d/%Y", "%Y-%m-%d"):
                        try:
                            return dt.datetime.strptime(m.group(1).strip(), fmt).date()
                        except ValueError:
                            continue
    return None


def _find_company(rows):
    """First non-empty text cell that isn't a report title / 'As of' line."""
    for row in rows[:4]:
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                low = cell.strip().lower()
                if low in ("balance sheet", "a/p aging summary", "ap aging summary") \
                        or low.startswith("as of"):
                    continue
                return cell.strip()
    return None


def _safe_filename(s):
    return re.sub(r'[\\/:*?"<>|]+', " ", s).strip() if s else s


# Master company-name mapping (QBO Company -> Short Name), loaded from the bundled
# Master_Mapping_List.xlsx: column A = QBO Company, column B = Short Name.
MAPPING_PATH = os.path.join(BASE_DIR, "Master_Mapping_List.xlsx")


def _company_key(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s else ""


def _load_company_map():
    out = {}
    if os.path.exists(MAPPING_PATH):
        try:
            wb = load_workbook(MAPPING_PATH, data_only=True, read_only=True)
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                qbo = row[0] if row and len(row) > 0 else None
                short = row[1] if row and len(row) > 1 else None
                if isinstance(qbo, str) and qbo.strip() and \
                        qbo.strip().upper() != "N/A" and short:
                    out[_company_key(qbo)] = str(short).strip()
            wb.close()
        except Exception:
            pass
    return out


COMPANY_MAP = _load_company_map()


def _short_name(*cands):
    for c in cands:
        key = _company_key(c)
        if key and key in COMPANY_MAP:
            return COMPANY_MAP[key]
    return None


def _first_text(row):
    for c in row:
        if isinstance(c, str) and c.strip():
            return c
    return None


def _norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""


def _norm_group(s):
    """Normalize a group label for exclusion/parent matching: drop a leading
    numbering prefix like '1-' and turn punctuation into spaces."""
    s = _norm(s)
    s = re.sub(r"^\d+\s*[-.)]\s*", "", s)        # '1-Missing Buy Ins' -> 'missing buy ins'
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# =========================================================================== #
# Name matching
# =========================================================================== #

def _alias_norm(name):
    s = _norm(name)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# Build alias groups: every member normalizes to a shared id.
_ALIAS_GROUP = {}
for _i, _pair in enumerate(ALIASES):
    for _n in _pair:
        _ALIAS_GROUP[_alias_norm(_n)] = f"__alias_{_i}__"


def _tokens(name):
    s = _norm(name)
    s = s.replace(".", "")                       # St. -> St, L.A. -> LA
    s = re.sub(r"[^a-z0-9]+", " ", s)
    toks = [t for t in s.split() if t]
    out = []
    for t in toks:
        out.extend(CITY_ABBREV[t].split() if t in CITY_ABBREV else [t])
    return out


def _key(name):
    """Return (core_token_frozenset, sport_or_None) used for comparison.
    SC/FC/CF are kept as tokens (set comparison already makes position irrelevant,
    e.g. 'FC Dallas' == 'Dallas FC'); dropping them caused false subset matches like
    'FC Dallas' -> 'Dallas Mavericks'."""
    core, sport = [], None
    for t in _tokens(name):
        if t in SPORTS:
            sport = t
        else:
            core.append(t)
    return frozenset(core), sport


def _sport_ok(a, b):
    return not (a and b and a != b)


def match_credits_to_ap(teams, ap_list):
    """Assign an A/P record to each ticket-credit team.

    Returns (team_ap_total, team_ap_idx, notes). Precision-first: ambiguous
    candidates are left unmatched. PSL A/P rows are never matched. `notes` records
    every non-exact match and every ambiguous / PSL skip for the Notes tab.
    """
    ap_keys = [_key(a["Vendor"]) for a in ap_list]
    ap_alias = [_ALIAS_GROUP.get(_alias_norm(a["Vendor"])) for a in ap_list]
    used = set()
    team_total = [0.0] * len(teams)
    team_idx = [-1] * len(teams)
    notes = []

    def candidates(tname):
        tcore, tsport = _key(tname)
        talias = _ALIAS_GROUP.get(_alias_norm(tname))
        alias_hits, exact_hits, subset_hits = [], [], []
        for j, a in enumerate(ap_list):
            if j in used or a.get("is_psl"):
                continue
            acore, asport = ap_keys[j]
            if talias and ap_alias[j] == talias:
                alias_hits.append(j)
                continue
            if not tcore or not acore or not _sport_ok(tsport, asport):
                continue
            if tcore == acore:
                exact_hits.append(j)
            elif tcore <= acore or acore <= tcore:
                subset_hits.append(j)
        for tier, hits in (("alias", alias_hits), ("exact", exact_hits),
                           ("subset", subset_hits)):
            if hits:
                return tier, hits
        return None, []

    for i, t in enumerate(teams):
        tier, hits = candidates(t["Vendor"])
        if len(hits) == 1:
            j = hits[0]
            used.add(j)
            team_idx[i] = j
            team_total[i] = ap_list[j]["Total"]
            if _alias_norm(t["Vendor"]) != _alias_norm(ap_list[j]["Vendor"]):
                notes.append(_match_note(tier, t["Vendor"], ap_list[j]["Vendor"]))
        elif len(hits) > 1:
            notes.append({
                "Category": "Not matched (ambiguous)",
                "Balance Sheet (TC)": t["Vendor"],
                "A/P Aging": ", ".join(ap_list[j]["Vendor"] for j in hits),
                "Explanation": "Several A/P vendors could match this credit, so it was "
                               "left unmatched to avoid a wrong offset. Add an alias to "
                               "force the intended one."})

    return team_total, team_idx, notes


def _raw_tokens(name):
    s = _norm(name).replace(".", "")
    return [t for t in re.sub(r"[^a-z0-9]+", " ", s).split() if t]


def _ordered_extra(extra, name):
    seen, out = set(), []
    for t in _tokens(name):
        if t in extra and t not in seen:
            seen.add(t)
            out.append(t)
    return " ".join(out)


def _match_note(tier, tc, ap):
    """Describe *why* a non-exact pair was matched, specific to this pair."""
    if tier == "alias":
        reason = "Matched using the alias list."
    else:
        tcc, tcs = _key(tc)
        apc, aps = _key(ap)
        bits = []
        raw = _raw_tokens(tc) + _raw_tokens(ap)
        for t in dict.fromkeys(raw):
            if t in CITY_ABBREV:
                exp = CITY_ABBREV[t]
                bits.append(f"read \u201c{t.upper()}\u201d as \u201c{exp.title()}\u201d")
        if (tcs or aps) and tcs != aps:
            bits.append(f"sport word \u201c{(tcs or aps).title()}\u201d is on only one side")
        extra_ap = _ordered_extra(apc - tcc, ap)
        extra_tc = _ordered_extra(tcc - apc, tc)
        if extra_ap:
            bits.append(f"A/P name adds \u201c{extra_ap}\u201d")
        if extra_tc:
            bits.append(f"Balance Sheet name adds \u201c{extra_tc}\u201d")
        if not bits:
            bits.append("names differ only in punctuation, spacing, or word order")
        reason = "Matched: " + "; ".join(bits) + "."
    return {"Category": "Non-exact match", "Balance Sheet (TC)": tc,
            "A/P Aging": ap, "Explanation": reason}


# =========================================================================== #
# Parsing
# =========================================================================== #

def parse_balance_sheet(rows):
    """Return (items, as_of). items is an ordered list scoped to the Prepaid
    Inventory / Ticket Credits subtree (Missing Buy Ins & Deposits excluded). A row
    is a 'header' only if it has child rows beneath it (deeper indent); leaf rows are
    'team' accounts whether or not they carry a '(TC)' suffix."""
    parent_pos, parent_ind = None, None
    for i, row in enumerate(rows):
        name = _first_text(row)
        if name is None:
            continue
        nm = name.strip()
        if _norm_group(nm) in PARENT_NAMES and not _norm(nm).startswith("total"):
            parent_pos = i
            parent_ind = len(name) - len(name.lstrip())
            break
    if parent_pos is None:
        raise ValueError("Could not find a 'Prepaid Inventory' (or 'Ticket Credits') "
                         "section in the Balance Sheet.")

    # capture every row of the section with its indent + amount
    section = []
    for row in rows[parent_pos + 1:]:
        name = _first_text(row)
        if name is None:
            continue
        ind = len(name) - len(name.lstrip())
        if ind <= parent_ind:
            break
        section.append((name.strip(), ind, _amount(row[-1]) if row else None))

    items = []
    skip_ind = None
    for i, (nm, ind, amt) in enumerate(section):
        if skip_ind is not None:
            if ind > skip_ind:
                continue                          # still inside excluded subgroup
            skip_ind = None
        if _norm(nm).startswith("total"):
            continue
        if _norm_group(nm) in EXCLUDED_GROUPS:
            skip_ind = ind
            continue
        next_ind = section[i + 1][1] if i + 1 < len(section) else parent_ind
        if next_ind > ind:                        # has children -> group header
            label = re.sub(r"\s*\(tc\)\s*$", "", nm, flags=re.I).strip()
            items.append({"kind": "header", "ind": ind, "Vendor": label})
        else:                                     # leaf account -> team
            if _norm(nm).endswith("(tc)"):
                vendor = re.sub(r"\s*\(tc\)\s*$", "", nm, flags=re.I).strip()
            else:
                vendor = nm
            items.append({"kind": "team", "ind": ind, "Vendor": vendor,
                          "Expense Account": nm, "Ticket Credit Amount": amt})
    return items, _find_asof(rows)


def extract_prepaid_section(rows):
    """Return [(account_name, amount), ...] for the Prepaid Inventory / Ticket
    Credits section exactly as it appears on the Balance Sheet (parent row through
    its matching Total line), for the source tab."""
    parent_pos, parent_ind = None, None
    for i, row in enumerate(rows):
        name = _first_text(row)
        if name is None:
            continue
        nm = name.strip()
        if _norm_group(nm) in PARENT_NAMES and not _norm(nm).startswith("total"):
            parent_pos, parent_ind = i, len(name) - len(name.lstrip())
            break
    if parent_pos is None:
        return []
    pr = _first_text(rows[parent_pos])
    out = [(pr.rstrip(), _amount(rows[parent_pos][-1]) if rows[parent_pos] else None)]
    for row in rows[parent_pos + 1:]:
        name = _first_text(row)
        if name is None:
            continue
        ind = len(name) - len(name.lstrip())
        amt = _amount(row[-1]) if row else None
        if ind <= parent_ind:
            if _norm(name.strip()).startswith("total"):
                out.append((name.rstrip(), amt))
            break
        out.append((name.rstrip(), amt))
    return out


_BUCKETS = ["Current", "1 - 30", "31 - 60", "61 - 90", "91 and over"]
_PSL_RE = re.compile(r"\bpsls?$")


def _is_psl(name):
    return bool(_PSL_RE.search(re.sub(r"[^a-z0-9]+", " ", _norm(name)).strip()))


def parse_ap_aging(rows):
    header_idx, col_map = None, {}
    for i, row in enumerate(rows[:15]):
        norm = [_norm(c) for c in row]
        if "current" in norm:
            header_idx = i
            for j, h in enumerate(norm):
                for key in _BUCKETS:
                    if h.replace(" ", "") == key.lower().replace(" ", ""):
                        col_map[key] = j
            break
    if header_idx is None:
        raise ValueError("Could not find the aging header row (no 'Current' column) "
                         "in the A/P Aging Summary.")
    missing = [k for k in _BUCKETS if k not in col_map]
    if missing:
        raise ValueError("A/P Aging Summary is missing column(s): " + ", ".join(missing))

    ap = []
    for row in rows[header_idx + 1:]:
        vname = row[0] if row else None
        if not isinstance(vname, str):
            continue
        v = vname.strip()
        if v == "" or v.upper() == "TOTAL" or "as of" in v.lower() or \
                re.match(r"^\s*\w+day,", v):
            continue
        buckets = []
        for key in _BUCKETS:
            j = col_map[key]
            val = _amount(row[j]) if j < len(row) else None
            buckets.append(val if val is not None else 0.0)
        ap.append({"Vendor": v, "buckets": buckets, "Total": round(sum(buckets), 2),
                   "is_psl": _is_psl(v)})
    return ap, _find_asof(rows)


# =========================================================================== #
# Reconciliation
# =========================================================================== #

def reconcile(items, ap_list, pay_date):
    teams = [it for it in items if it["kind"] == "team"]
    team_total, team_idx, notes = match_credits_to_ap(teams, ap_list)

    # attach comparison results to each team (index aligned with `teams`)
    ti = 0
    for it in items:
        if it["kind"] != "team":
            continue
        TC = it.get("Ticket Credit Amount") or 0.0
        AP = team_total[ti]
        it["_ap_idx"] = team_idx[ti]
        if TC < AP:
            opt, lower = "TC", TC
        elif AP < TC:
            opt, lower = "AP", AP
        else:
            opt, lower = "Equal", TC
        it.update({"Ticket Credit Amount": TC, "AP Aging Amount": AP,
                   "Lower Option": opt, "Lower Amount": round(lower, 2),
                   "Difference (AP - TC)": round(AP - TC, 2),
                   "Payment Date": pay_date})
        ti += 1

    def vc_keep(it):
        return (it["Lower Amount"] != 0 or it["Lower Option"] == "Equal") \
            and it["Lower Amount"] >= 0

    def otc_keep(it):
        return it["Ticket Credit Amount"] >= 0 and it["AP Aging Amount"] <= 0

    # Build ordered Vendor Credits / Other TC. A surviving team pulls in all of
    # its not-yet-emitted ancestor league headers (so NCAA > Football nests), and
    # a header with no surviving descendants in that section is omitted.
    def assemble(keep):
        out, stack = [], []          # stack: [[ind, header_item, emitted], ...]
        for it in items:
            if it["kind"] == "header":
                while stack and stack[-1][0] >= it["ind"]:
                    stack.pop()
                stack.append([it["ind"], it, False])
            elif keep(it):
                for entry in stack:
                    if not entry[2]:
                        out.append(entry[1])
                        entry[2] = True
                out.append(it)
        return out

    vc = assemble(vc_keep)
    otc = assemble(otc_keep)

    # Other AP: A/P rows whose vendor was NOT offset by a surviving Vendor Credit.
    offset_ap = {it["_ap_idx"] for it in items
                 if it["kind"] == "team" and vc_keep(it) and it["_ap_idx"] >= 0}
    other_ap = [a for j, a in enumerate(ap_list) if j not in offset_ap]
    return vc, otc, other_ap, notes


# =========================================================================== #
# Workbook builder
# =========================================================================== #

CUR = '$#,##0.00;-$#,##0.00'
DATEFMT = "m/d/yyyy"
HDR_FILL = PatternFill("solid", fgColor="5B9BD5")          # blue column header
BAND_FILL = PatternFill("solid", fgColor="DDEBF7")         # light blue banding
COL1_FILL = PatternFill("solid", fgColor="D9D9D9")         # light gray middle spacer
CENTER = Alignment(horizontal="center", vertical="center")
ARIAL = "Arial"


def _autofit(ws, columns):
    for ci, col in enumerate(columns, start=1):
        if col == "Column1":
            ws.column_dimensions[get_column_letter(ci)].width = 3
            continue
        maxlen = (len(str(col)) + 3) if col else 0   # +3 reserves the filter arrow
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, ci).value
            if v is None:
                continue
            if isinstance(v, dt.datetime):
                ln = 10
            elif isinstance(v, (int, float)):
                nf = ws.cell(r, ci).number_format or ""
                ln = len(f"{abs(v):,.2f}") + (2 if "$" in nf else 0) + (1 if v < 0 else 0)
            else:
                ln = len(str(v))
            maxlen = max(maxlen, ln)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 2, 6), 60)


def _emit_sheet(wb, title, columns, records, *, money_cols, blankzero_cols,
                col1_present):
    """The three output tabs: blue header, banded, centered, league headers bold in
    column A only, blank 'Column1' header label, gray Column1 spacer, no totals."""
    ws = wb.create_sheet(title)
    idx = {c: i for i, c in enumerate(columns)}

    ws.append(["" if c == "Column1" else c for c in columns])
    for c in range(1, len(columns) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 18

    for n, rec in enumerate(records):
        is_header = rec.get("kind") == "header"
        band = BAND_FILL if n % 2 == 1 else None
        rowvals = []
        for col in columns:
            v = rec.get(col, "")
            if col in money_cols and col in blankzero_cols and v in (0, 0.0, None, ""):
                v = None
            if isinstance(v, dt.date):
                v = dt.datetime(v.year, v.month, v.day)
            rowvals.append(v)
        ws.append(rowvals)
        r = ws.max_row
        for c in range(1, len(columns) + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=ARIAL, size=10, bold=(is_header and c == 1))
            cell.alignment = CENTER
            if band:
                cell.fill = band
            col = columns[c - 1]
            if col in money_cols:
                cell.number_format = CUR
            elif col == "Payment Date":
                cell.number_format = DATEFMT
        if col1_present and "Column1" in idx:
            ws.cell(r, idx["Column1"] + 1).fill = COL1_FILL

    _autofit(ws, columns)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"
    return ws


def _emit_table(wb, title, columns, rows, *, money_cols=(), left_cols=()):
    """Source / Notes tabs: blue header, banding, autofit. Text columns left-aligned,
    numeric columns right-aligned, no league headers or Column1 spacer."""
    ws = wb.create_sheet(title)
    ws.append(list(columns))
    for c in range(1, len(columns) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 18

    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    right = Alignment(horizontal="right", vertical="center")
    for n, row in enumerate(rows):
        vals = []
        for col in columns:
            v = row.get(col, "") if isinstance(row, dict) else row[columns.index(col)]
            if isinstance(v, dt.date):
                v = dt.datetime(v.year, v.month, v.day)
            vals.append(v)
        ws.append(vals)
        r = ws.max_row
        band = BAND_FILL if n % 2 == 1 else None
        for c in range(1, len(columns) + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=ARIAL, size=10)
            col = columns[c - 1]
            if col in money_cols:
                cell.number_format = CUR
                cell.alignment = right
            else:
                cell.alignment = left
            if band:
                cell.fill = band
    _autofit(ws, columns)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"
    return ws


def build_workbook(bs_files, ap_files, override_date):
    bs_rows = []
    for fn, data in bs_files:
        bs_rows.extend(_read_balance_sheet(fn, data))
    ap_rows = []
    for fn, data in ap_files:
        ap_rows.extend(_rows_from_upload(fn, data))

    items, asof_bs = parse_balance_sheet(bs_rows)
    teams = [it for it in items if it["kind"] == "team"]
    if not teams:
        raise ValueError("No ticket-credit (leaf) accounts were found in the Prepaid "
                         "Inventory section of the Balance Sheet.")
    ap_list, asof_ap = parse_ap_aging(ap_rows)
    if not ap_list:
        raise ValueError("No vendor rows found in the A/P Aging Summary.")

    company_bs = _find_company(bs_rows)
    company_ap = _find_company(ap_rows)
    company = company_ap or company_bs or "Company"
    company_short = _short_name(company_ap, company_bs)

    # Date checks: the two reports should share the same date, and it should be a
    # calendar month-end. These produce warnings (not errors).
    warnings_list = []

    def _is_month_end(d):
        return bool(d) and d.day == calendar.monthrange(d.year, d.month)[1]

    if asof_bs and asof_ap and asof_bs != asof_ap:
        warnings_list.append(
            f"The Balance Sheet is dated {asof_bs:%b %-d, %Y} but the A/P Aging is dated "
            f"{asof_ap:%b %-d, %Y}. Both reports should use the same date.")
    for label, d in (("Balance Sheet", asof_bs), ("A/P Aging", asof_ap)):
        if d and not _is_month_end(d):
            warnings_list.append(
                f"The {label} date ({d:%b %-d, %Y}) is not the last day of the month.")

    pay_date = override_date or asof_bs or asof_ap or dt.date.today()
    vc, otc, other_ap, notes = reconcile(items, ap_list, pay_date)

    def header_rec(it):                 # league header: only the name in column A
        return {"kind": "header", "Vendor": it["Vendor"]}

    # ---- Vendor Credits records ----
    vc_cols = ["Vendor", "Payment Date", "Expense Account", "Expense Line Amount",
               "Column1", "Lower Option", "Ticket Credit Amount", "AP Aging Amount",
               "Difference (AP - TC)"]
    vc_recs = []
    for it in vc:
        if it["kind"] == "header":
            vc_recs.append(header_rec(it))
        else:
            vc_recs.append({"kind": "team", "Vendor": it["Vendor"],
                            "Payment Date": pay_date,
                            "Expense Account": it["Expense Account"],
                            "Expense Line Amount": it["Lower Amount"], "Column1": "",
                            "Lower Option": it["Lower Option"],
                            "Ticket Credit Amount": it["Ticket Credit Amount"],
                            "AP Aging Amount": it["AP Aging Amount"],
                            "Difference (AP - TC)": it["Difference (AP - TC)"]})
    vc_money = {"Expense Line Amount", "Ticket Credit Amount", "AP Aging Amount",
                "Difference (AP - TC)"}

    def add_vendor_credits(target_wb, cols, col1):
        _emit_sheet(target_wb, "Vendor Credits", cols, vc_recs,
                    money_cols=vc_money & set(cols),
                    blankzero_cols={"Expense Line Amount"}, col1_present=col1)

    # ---- combined workbook ----
    wb = Workbook()
    wb.remove(wb.active)
    add_vendor_credits(wb, vc_cols, True)

    # Other TC
    otc_cols = ["Vendor", "Payment Date", "Expense Account", "Ticket Credit Amount"]
    otc_recs = []
    for it in otc:
        if it["kind"] == "header":
            otc_recs.append(header_rec(it))
        else:
            otc_recs.append({"kind": "team", "Vendor": it["Vendor"],
                             "Payment Date": pay_date,
                             "Expense Account": it["Expense Account"],
                             "Ticket Credit Amount": it["Ticket Credit Amount"]})
    _emit_sheet(wb, "Other TC", otc_cols, otc_recs,
                money_cols={"Ticket Credit Amount"},
                blankzero_cols={"Ticket Credit Amount"}, col1_present=False)

    # Other AP
    oap_cols = ["Vendor", "Payment Date", "Expense Account", "Expense Line Amount",
                "Column1", "Current", "1 - 30", "31 - 60", "61 - 90", "91 and over"]
    oap_recs = []
    for a in other_ap:
        rec = {"kind": "team", "Vendor": a["Vendor"], "Payment Date": pay_date,
               "Expense Account": "Cost of Goods Sold",
               "Expense Line Amount": a["Total"], "Column1": ""}
        for key, val in zip(_BUCKETS, a["buckets"]):
            rec[key] = val
        oap_recs.append(rec)
    _emit_sheet(wb, "Other AP", oap_cols, oap_recs,
                money_cols={"Expense Line Amount", "Current", "1 - 30", "31 - 60",
                            "61 - 90", "91 and over"},
                blankzero_cols={"Current", "1 - 30", "31 - 60", "61 - 90", "91 and over"},
                col1_present=True)

    # ---- source tab: Ticket Credits (Prepaid Inventory section only) ----
    section = extract_prepaid_section(bs_rows)
    tc_src_rows = [{"Account": nm, "Amount": amt} for nm, amt in section]
    _emit_table(wb, "Ticket Credits", ["Account", "Amount"], tc_src_rows,
                money_cols={"Amount"}, left_cols={"Account"})

    # ---- source tab: A/P Aging (as uploaded) ----
    ap_cols = ["Vendor", "Current", "1 - 30", "31 - 60", "61 - 90", "91 and over", "Total"]
    ap_src_rows = []
    for a in ap_list:
        row = {"Vendor": a["Vendor"], "Total": a["Total"]}
        for key, val in zip(_BUCKETS, a["buckets"]):
            row[key] = val
        ap_src_rows.append(row)
    _emit_table(wb, "AP Aging", ap_cols, ap_src_rows,
                money_cols={"Current", "1 - 30", "31 - 60", "61 - 90", "91 and over",
                            "Total"})

    # ---- Notes ----
    note_cols = ["Category", "Balance Sheet (TC)", "A/P Aging", "Explanation"]
    note_rows = notes if notes else [{
        "Category": "All matches exact", "Balance Sheet (TC)": "", "A/P Aging": "",
        "Explanation": "Every ticket-credit account matched an A/P vendor by exact name "
                       "(after case/spacing normalization), or had no A/P counterpart."}]
    _emit_table(wb, "Notes", note_cols, note_rows)

    buf = io.BytesIO()
    wb.save(buf)

    # ---- VC-only workbook (for QBO upload): columns A-D only ----
    wb_vc = Workbook()
    wb_vc.remove(wb_vc.active)
    add_vendor_credits(wb_vc, vc_cols[:4], False)
    buf_vc = io.BytesIO()
    wb_vc.save(buf_vc)

    n_team_vc = sum(1 for it in vc if it["kind"] == "team")
    credit_applied = round(sum(it["Lower Amount"] for it in vc if it["kind"] == "team"), 2)
    meta = {
        "company": company,
        "company_short": company_short,
        "as_of": pay_date.strftime("%B %d, %Y"),
        "tc_accounts": len(teams),
        "ap_vendors": len(ap_list),
        "vendor_credits_rows": n_team_vc,
        "credit_applied": credit_applied,
        "other_tc_rows": sum(1 for it in otc if it["kind"] == "team"),
        "other_tc_total": round(sum(it["Ticket Credit Amount"] for it in otc
                                    if it["kind"] == "team"), 2),
        "other_ap_rows": len(other_ap),
        "other_ap_total": round(sum(a["Total"] for a in other_ap), 2),
        "tc_total": round(sum((t.get("Ticket Credit Amount") or 0.0) for t in teams), 2),
        "ap_total": round(sum(a["Total"] for a in ap_list), 2),
        "notes_count": len(notes),
        "warnings": warnings_list,
    }
    return buf.getvalue(), buf_vc.getvalue(), meta


# =========================================================================== #
# Routes
# =========================================================================== #

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/process", methods=["POST"])
def process():
    bs = [(f.filename, f.read()) for f in request.files.getlist("balance_sheet") if f.filename]
    ap = [(f.filename, f.read()) for f in request.files.getlist("ap_aging") if f.filename]
    if not bs:
        return jsonify({"error": "Please upload a Balance Sheet."}), 400
    if not ap:
        return jsonify({"error": "Please upload an A/P Aging Summary."}), 400

    override = None
    raw = (request.form.get("month_end") or "").strip()
    if raw:
        try:
            override = dt.datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "Month-end date must be YYYY-MM-DD."}), 400

    try:
        data, data_vc, meta = build_workbook(bs, ap, override)
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    token = uuid.uuid4().hex
    base = _safe_filename(meta.get("company_short") or meta["company"]) or "Company"
    fn_all = f"{base} - Vendor Credits {meta['as_of']}.xlsx"
    fn_vc = f"{base} - Vendor Credits (QBO) {meta['as_of']}.xlsx"
    folder = os.path.join(STORE_DIR, token)
    os.makedirs(folder, exist_ok=True)
    with open(os.path.join(folder, fn_all), "wb") as fh:
        fh.write(data)
    with open(os.path.join(folder, fn_vc), "wb") as fh:
        fh.write(data_vc)
    _cleanup_old()
    meta["download_url"] = f"/download/{token}"
    meta["download_vc_url"] = f"/download/{token}?which=vc"
    meta["filename"] = fn_all
    meta["filename_vc"] = fn_vc
    return jsonify(meta)


@app.route("/download/<token>")
def download(token):
    folder = os.path.join(STORE_DIR, os.path.basename(token))
    if not os.path.isdir(folder):
        abort(404)
    xlsx = [f for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    if not xlsx:
        abort(404)
    want_vc = request.args.get("which") == "vc"
    vc_files = [f for f in xlsx if "(QBO)" in f]
    if want_vc:
        pick = vc_files[0] if vc_files else xlsx[0]
    else:
        others = [f for f in xlsx if "(QBO)" not in f]
        pick = others[0] if others else xlsx[0]
    return send_file(os.path.join(folder, pick),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=pick)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
